# =============================================================
#  Form 26AS Reconciliation Utility — Flask Backend
#  Run:  python app.py
#  Then open:  http://localhost:5000
# =============================================================

import os
import io
import math
import csv
import uuid
from datetime import datetime
from openpyxl import load_workbook
from flask import Flask, jsonify, request, send_file, render_template
from flask_cors import CORS
import mysql.connector
from mysql.connector import Error as MySQLError
from dotenv import load_dotenv
from pypdf import PdfReader
import hashlib
import re

load_dotenv()

# ── BigQuery sync (optional — gracefully skips if not configured) ──
try:
    from google.cloud import bigquery
    BQ_CLIENT = bigquery.Client.from_service_account_json(
        os.path.join(os.path.dirname(__file__), 'bq_service_account.json')
    )
    BQ_PROJECT = "eternal-skyline-468312-g7"
    BQ_DATASET = "tds_reconciliation"
    BQ_ENABLED = True
    print("[BQ] BigQuery sync enabled")
except Exception as _bq_err:
    BQ_CLIENT = None
    BQ_ENABLED = False
    print(f"[BQ] BigQuery sync disabled: {_bq_err}")

def bq_upsert_rows(table_id, rows):
    """Insert/replace rows into a BigQuery table. Creates table if needed."""
    if not BQ_ENABLED or not rows:
        return
    try:
        full_table = f"{BQ_PROJECT}.{BQ_DATASET}.{table_id}"
        errors = BQ_CLIENT.insert_rows_json(full_table, rows)
        if errors:
            print(f"[BQ] Insert errors for {table_id}: {errors}")
    except Exception as e:
        print(f"[BQ] Sync error for {table_id}: {e}")

def bq_sync_income_records(fin_year, cust_ids=None):
    """Sync income_records for a given FY (and optionally specific cust_ids) to BigQuery."""
    if not BQ_ENABLED:
        return
    try:
        conn = get_db()
        cur = conn.cursor(dictionary=True)
        if cust_ids:
            placeholders = ','.join(['%s'] * len(cust_ids))
            cur.execute(f"""
                SELECT ir.income_id, ir.upload_ref, ir.fin_year, ir.cust_id,
                       cm.cust_code, cm.cust_name, cm.pan_number, cm.tan_number,
                       ir.doc_number, CAST(ir.doc_date AS CHAR) AS doc_date,
                       ir.doc_type, ir.taxable_amt, ir.tax_amt, ir.gross_amt,
                       ir.tds_section, ir.tds_rate, ir.entry_source
                FROM income_records ir
                JOIN cust_master cm ON ir.cust_id = cm.cust_id
                WHERE ir.fin_year = %s AND ir.cust_id IN ({placeholders})
            """, [fin_year] + list(cust_ids))
        else:
            cur.execute("""
                SELECT ir.income_id, ir.upload_ref, ir.fin_year, ir.cust_id,
                       cm.cust_code, cm.cust_name, cm.pan_number, cm.tan_number,
                       ir.doc_number, CAST(ir.doc_date AS CHAR) AS doc_date,
                       ir.doc_type, ir.taxable_amt, ir.tax_amt, ir.gross_amt,
                       ir.tds_section, ir.tds_rate, ir.entry_source
                FROM income_records ir
                JOIN cust_master cm ON ir.cust_id = cm.cust_id
                WHERE ir.fin_year = %s
            """, (fin_year,))
        rows = cur.fetchall()
        _close(cur, conn)
        # Convert Decimal to float for BQ JSON serialisation
        clean = []
        for row in rows:
            clean.append({k: float(v) if hasattr(v, '__float__') and not isinstance(v, (int, str, type(None))) else
                           (str(v) if not isinstance(v, (int, float, bool, type(None))) else v)
                          for k, v in row.items()})
        full_table = f"{BQ_PROJECT}.{BQ_DATASET}.income_records"
        errors = BQ_CLIENT.insert_rows_json(full_table, clean)
        if errors:
            print(f"[BQ] income_records insert errors: {errors}")
        else:
            print(f"[BQ] Synced {len(clean)} income_records rows for FY {fin_year}")
    except Exception as e:
        print(f"[BQ] bq_sync_income_records error: {e}")

def bq_sync_form26as(fin_year):
    """Sync form26as_records for a given FY to BigQuery."""
    if not BQ_ENABLED:
        return
    try:
        conn = get_db()
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT f.record_id, f.fin_year, f.tan_of_deductor, f.deductor_name,
                   f.section_code, CAST(f.transaction_date AS CHAR) AS transaction_date,
                   f.booking_amt, f.tds_credited,
                   cm.cust_code, cm.cust_name
            FROM form26as_records f
            LEFT JOIN cust_master cm ON LOWER(TRIM(cm.tan_number)) = LOWER(TRIM(f.tan_of_deductor))
            WHERE f.fin_year = %s
        """, (fin_year,))
        rows = cur.fetchall()
        _close(cur, conn)
        clean = []
        for row in rows:
            clean.append({k: float(v) if hasattr(v, '__float__') and not isinstance(v, (int, str, type(None))) else
                           (str(v) if not isinstance(v, (int, float, bool, type(None))) else v)
                          for k, v in row.items()})
        full_table = f"{BQ_PROJECT}.{BQ_DATASET}.form26as_records"
        errors = BQ_CLIENT.insert_rows_json(full_table, clean)
        if errors:
            print(f"[BQ] form26as_records insert errors: {errors}")
        else:
            print(f"[BQ] Synced {len(clean)} form26as_records rows for FY {fin_year}")
    except Exception as e:
        print(f"[BQ] bq_sync_form26as error: {e}")


def update_tds_deductible_from_section(fin_year):
    """
    After Form 26AS upload, updates income_records.tds_deductible based on 
    annual cumulative thresholds and legally compounded tax rate logic.
    Also automatically flags non-PAN baseline fallback rate increases.
    """
    if not fin_year:
        return
    try:
        conn = get_db()
        cur = conn.cursor(dictionary=True)

        # Step 1 — Get ALL valid distinct sections per customer instead of an arbitrary MAX()
        cur.execute("""
            SELECT DISTINCT
                cm.cust_id,
                cm.pan_number,
                f.section_code
            FROM cust_master cm
            JOIN form26as_records f 
              ON LOWER(TRIM(cm.tan_number)) = LOWER(TRIM(f.tan_of_deductor))
            WHERE f.fin_year = %s
              AND f.section_code IS NOT NULL
              AND f.section_code NOT IN ('', '—', '-')
        """, (fin_year,))
        customer_sections = cur.fetchall()

        updated = 0
        PAN_RE = re.compile(r'^[A-Z]{5}[0-9]{4}[A-Z]$')
        
        for entry in customer_sections:
            cust_id = entry['cust_id']
            pan_val = str(entry['pan_number'] or '').strip().upper()
            section_code = entry['section_code'].strip().upper()

            # Verify PAN validity state to detect penalty parameters
            has_valid_pan = bool(PAN_RE.match(pan_val))

            # Step 2 — Pull global section parameters
            cur.execute("""
                SELECT
                    tds_rate, surcharge_pct, cess_pct, threshold_limit, payee_type
                FROM tds_codes_master
                WHERE UPPER(TRIM(section_code)) = %s
                  AND is_resident = 1
                  AND is_active = 1
                ORDER BY
                    CASE payee_type
                        WHEN 'OTHERS'         THEN 1
                        WHEN 'ALL'            THEN 2
                        WHEN 'INDIVIDUAL_HUF' THEN 3
                        ELSE 4
                    END
                LIMIT 1
            """, (section_code,))
            rate_row = cur.fetchone()
            
            if not rate_row:
                continue

            base_rate = float(rate_row['tds_rate'] or 0.0)
            surcharge = float(rate_row['surcharge_pct'] or 0.0)
            cess      = float(rate_row['cess_pct'] or 0.0)
            threshold = rate_row['threshold_limit'] 

            # STATUTORY PRODUCTION CHECK 1: NON-PAN EXCEPTION PENALTY LAWS
            # If PAN missing/invalid, rate spikes to higher of section rule vs 20.00%
            if not has_valid_pan:
                # Sections like 194B/194BA carry a baseline of 30%, so evaluate max bounds
                base_rate = max(base_rate, 20.00)

            # STATUTORY PRODUCTION CHECK 2: CORRECT COMPOUND MATH FORMULA
            # Effective Rate = Base Rate * (1 + Surcharge%) * (1 + Cess%)
            effective_rate = base_rate * (1 + (surcharge / 100)) * (1 + (cess / 100))

            # Step 3 — Calculate total annual cumulative aggregate taxable amount for threshold validation
            cur.execute("""
                SELECT COALESCE(SUM(taxable_amt), 0) AS total_annual_taxable
                FROM income_records
                WHERE cust_id = %s AND fin_year = %s AND UPPER(TRIM(tds_section)) = %s
            """, (cust_id, fin_year, section_code))
            annual_summary = cur.fetchone()
            annual_taxable_total = float(annual_summary['total_annual_taxable'] if annual_summary else 0.0)

            # Step 4 — Fetch individual lines for THIS section only
            # Without section filter a customer with 194C+194J rows would have
            # all rows overwritten with whichever section loops last
            cur.execute("""
                SELECT income_id, taxable_amt, tax_amt
                FROM income_records
                WHERE cust_id = %s AND fin_year = %s AND UPPER(TRIM(tds_section)) = %s
            """, (cust_id, fin_year, section_code))
            invoices = cur.fetchall()

            for inv in invoices:
                income_id = inv['income_id']
                taxable_amt = float(inv['taxable_amt'] or 0)
                tax_amt = float(inv['tax_amt'] or 0)

                # Check if the annual aggregate total fails the benchmark rule limits
                if threshold and annual_taxable_total < threshold:
                    tds_deductible = 0.00
                elif effective_rate > 0:
                    tds_deductible = round(taxable_amt * effective_rate / 100, 2)
                else:
                    tds_deductible = tax_amt  # Fallback to base ledger values

                cur.execute("""
                    UPDATE income_records
                    SET tds_deductible = %s,
                        tds_section    = %s
                    WHERE income_id = %s
                """, (tds_deductible, section_code, income_id))
                updated += cur.rowcount

        conn.commit()
        _close(cur, conn)
        print(f"[TDS] Updated tds_deductible for {updated} rows successfully for FY {fin_year}")

    except Exception as e:
        print(f"[TDS] update_tds_deductible_from_section error: {e}")


def bq_sync_tds_summary(fin_year):
    """Compute TDS summary from MySQL and sync snapshot to BigQuery."""
    if not BQ_ENABLED:
        return
    try:
        conn = get_db()
        conn.commit()
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT
                c.cust_id,
                c.cust_code,
                c.cust_name,
                c.pan_number,
                c.tan_number,
                ir.fin_year,
                COALESCE(SUM(ir.taxable_amt), 0.00)    AS total_taxable_amt,
                COALESCE(SUM(ir.tds_deductible), 0.00)  AS total_tds_deductible,
                COALESCE(SUM(ir.gross_amt), 0.00)      AS total_gross_amt,
                COALESCE(f26.total_tds_credited, 0.00) AS total_26as_credit,
                COALESCE(SUM(ir.tds_deductible), 0.00)
                    - COALESCE(f26.total_tds_credited, 0.00) AS variance_amt,
                CASE WHEN COALESCE(SUM(ir.tds_deductible), 0) = 0 THEN 0.00
                     ELSE ROUND(COALESCE(f26.total_tds_credited, 0.00)
                          / COALESCE(SUM(ir.tds_deductible), 0.00) * 100, 2)
                END AS pct_credit,
                COUNT(ir.income_id) AS invoice_count
            FROM cust_master c
            LEFT JOIN income_records ir ON c.cust_id = ir.cust_id AND ir.fin_year = %s
            LEFT JOIN (
                SELECT LOWER(TRIM(tan_of_deductor)) AS tan, SUM(tds_credited) AS total_tds_credited
                FROM form26as_records WHERE fin_year = %s
                GROUP BY LOWER(TRIM(tan_of_deductor))
            ) f26 ON LOWER(TRIM(c.tan_number)) = f26.tan
            WHERE ir.fin_year = %s
            GROUP BY c.cust_id, c.cust_code, c.cust_name, c.pan_number,
                     c.tan_number, ir.fin_year, f26.total_tds_credited
        """, (fin_year, fin_year, fin_year))
        rows = cur.fetchall()
        _close(cur, conn)
        if not rows:
            print(f"[BQ] bq_sync_tds_summary: no rows for FY {fin_year}")
            return
        clean = []
        for row in rows:
            clean.append({k: float(v) if hasattr(v, '__float__') and not isinstance(v, (int, str, type(None)))
                          else (str(v) if not isinstance(v, (int, float, bool, type(None))) else v)
                          for k, v in row.items()})
        full_table = f"{BQ_PROJECT}.{BQ_DATASET}.tds_summary"
        errors = BQ_CLIENT.insert_rows_json(full_table, clean)
        if errors:
            print(f"[BQ] tds_summary insert errors: {errors}")
        else:
            print(f"[BQ] Synced {len(clean)} tds_summary rows for FY {fin_year}")
    except Exception as e:
        print(f"[BQ] bq_sync_tds_summary error: {e}")


def bq_sync_cust_master():
    """Sync full cust_master to BigQuery using a fresh DB connection."""
    if not BQ_ENABLED:
        return
    try:
        conn = get_db()
        conn.commit()  # flush any pending transaction state
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT cust_id, cust_code, cust_name, pan_number, gstin_number, tan_number, cust_status FROM cust_master")
        rows = cur.fetchall()
        _close(cur, conn)
        if not rows:
            print("[BQ] bq_sync_cust_master: no rows found in MySQL")
            return
        # Ensure all values are BQ-serialisable
        clean = []
        for row in rows:
            clean.append({k: (str(v) if v is not None else None) if not isinstance(v, (int, float, bool, type(None))) else v
                          for k, v in row.items()})
        full_table = f"{BQ_PROJECT}.{BQ_DATASET}.cust_master"
        errors = BQ_CLIENT.insert_rows_json(full_table, clean)
        if errors:
            print(f"[BQ] cust_master insert errors: {errors}")
        else:
            print(f"[BQ] Synced {len(clean)} cust_master rows")
    except Exception as e:
        print(f"[BQ] bq_sync_cust_master error: {e}")


app = Flask(__name__)
CORS(app, resources={
    r"/api/*": {
        "origins": ["http://127.0.0.1:5000", "http://localhost:5000"],
        "methods": ["GET", "POST", "PUT", "DELETE"],
        "allow_headers": ["Content-Type", "X-Requested-With"]
    }
})


# ══════════════════════════════════════════════════════════════
#  DATABASE
# ══════════════════════════════════════════════════════════════

def get_db():
    try:
        return mysql.connector.connect(
            host      = os.getenv('DB_HOST',     'localhost'),
            port      = int(os.getenv('DB_PORT', '3306')),
            user      = os.getenv('DB_USER',     ''),
            password  = os.getenv('DB_PASSWORD', ''),
            database  = os.getenv('DB_NAME',     'form26as_db'),
            charset   = 'utf8mb4',
            autocommit= True,
        )
    except MySQLError as err:
        raise RuntimeError(f'Database connection failed: {err}')

def extract_fy_from_date_string(date_val):
    if not date_val:
        return None
    from datetime import datetime
    import pandas as pd
    
    if isinstance(date_val, (datetime, pd.Timestamp)):
        dt = date_val
    else:
        date_str = str(date_val).strip()
        if not date_str or date_str.lower() in ('none', 'nat', 'nan'):
            return None
        dt = None
        for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y', '%Y-%m-%d %H:%M:%S', '%d-%b-%Y'):
            try:
                dt = datetime.strptime(date_str, fmt)
                break
            except ValueError:
                continue
                
    if dt:
        fy_start = dt.year if dt.month >= 4 else dt.year - 1
        return f"FY {fy_start}-{str(fy_start + 1)[2:]}"
    return None



def _close(cursor, conn):
    for obj in (cursor, conn):
        try:
            obj.close()
        except Exception:
            pass


def _error(msg, status=400):
    return jsonify({'error': msg}), status




# ═════════════════════════════════════════════════════════════
# PAGE ROUTES — serve HTML files
# ═════════════════════════════════════════════════════════════

@app.route('/')
def index():
    return render_template('customer_master.html')
@app.route('/customer-master')
def page_customer_master():
    return render_template('customer_master.html')

@app.route('/form26as')
def page_form26as():
    return render_template('form26as.html')

@app.route('/books-of-accounts')
def page_books():
    return render_template('books_of_accounts.html')

@app.route('/upload-utility')
def page_upload_utility():
    return render_template('upload_utility.html')

@app.route('/reconciliation')
def page_reconciliation():
    return render_template('reconciliation.html')

@app.route('/tds-summary')
def page_tds_summary():
    return render_template('tds_summary.html')


@app.route('/audit-trail')
def page_audit_trail():
    # Renders the audit_trail.html template layout cleanly on the server
    return render_template('audit-trail.html')


# ══════════════════════════════════════════════════════════════
#  API — CUSTOMER MASTER
# ══════════════════════════════════════════════════════════════

# ── Stats (widget cards) ──────────────────────────────────────
@app.route('/api/customers/stats', methods=['GET'])
def customer_stats():
    try:
        conn   = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute('SELECT * FROM vw_cust_master_stats')
        row = cursor.fetchone()
        _close(cursor, conn)
        return jsonify({
            'total_customers'   : int(row['total_customers']    or 0),
            'active_customers'  : int(row['active_customers']   or 0),
            'inactive_customers': int(row['inactive_customers']  or 0),
            'total_tans'        : int(row['total_tans']          or 0),
        })
    except RuntimeError as e:
        return _error(str(e), 503)
    except MySQLError as e:
        return _error(f'Database error: {e}', 500)


# ── List customers ────────────────────────────────────────────
#   GET /api/customers?search=&status=&page=&limit=
@app.route('/api/customers', methods=['GET'])
def list_customers():
    search = request.args.get('search', '').strip()
    status = request.args.get('status', '').strip().lower()
    try:
        page  = max(1, int(request.args.get('page',  1)))
        limit = min(500, max(1, int(request.args.get('limit', 50))))
    except ValueError:
        return _error('page and limit must be integers')

    clauses, params = [], []

    if search:
        like = f'%{search}%'
        clauses.append(
            '(cust_name LIKE %s OR pan_number LIKE %s '
            'OR gstin_number LIKE %s OR tan_number LIKE %s OR cust_code LIKE %s)'
        )
        params.extend([like, like, like, like, like])

    if status in ('active', 'inactive'):
        clauses.append('cust_status = %s')
        params.append(status)

    where = ('WHERE ' + ' AND '.join(clauses)) if clauses else ''
    offset = (page - 1) * limit

    try:
        conn   = get_db()
        cursor = conn.cursor(dictionary=True)

        cursor.execute(f'SELECT COUNT(*) AS cnt FROM cust_master {where}', params)
        total = cursor.fetchone()['cnt']

        cursor.execute(
            f'''SELECT cust_id, cust_code, cust_name,
                       pan_number, gstin_number, tan_number,
                       cust_status, contact
                FROM cust_master {where}
                ORDER BY cust_name ASC
                LIMIT %s OFFSET %s''',
            params + [limit, offset]
        )
        rows = cursor.fetchall()
        _close(cursor, conn)

        return jsonify({
            'data' : rows,
            'total': total,
            'page' : page,
            'pages': math.ceil(total / limit) if total else 1,
            'limit': limit,
        })
    except RuntimeError as e:
        return _error(str(e), 503)
    except MySQLError as e:
        return _error(f'Database error: {e}', 500)


# ── Single customer ───────────────────────────────────────────
@app.route('/api/customers/<int:cust_id>', methods=['GET'])
def get_customer(cust_id):
    try:
        conn   = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            '''SELECT cust_id, cust_code, cust_name,
                      pan_number, gstin_number, tan_number,
                      cust_status, contact
               FROM cust_master WHERE cust_id = %s''',
            (cust_id,)
        )
        row = cursor.fetchone()
        _close(cursor, conn)
        if not row:
            return _error('Customer not found', 404)
        return jsonify(row)
    except RuntimeError as e:
        return _error(str(e), 503)
    except MySQLError as e:
        return _error(f'Database error: {e}', 500)

# ── Export CSV ────────────────────────────────────────────────
@app.route('/api/customers/export', methods=['GET'])
def export_customers():
    search = request.args.get('search', '').strip()
    status = request.args.get('status', '').strip().lower()
    clauses, params = [], []

    if search:
        like = f'%{search}%'
        clauses.append(
            '(cust_name LIKE %s OR pan_number LIKE %s '
            'OR gstin_number LIKE %s OR tan_number LIKE %s OR cust_code LIKE %s)'
        )
        params.extend([like, like, like, like, like])

    if status in ('active', 'inactive'):
        clauses.append('cust_status = %s')
        params.append(status)

    where = ('WHERE ' + ' AND '.join(clauses)) if clauses else ''

    try:
        conn   = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            f'''SELECT cust_code AS "Customer Code", cust_name AS "Customer Name",
                       pan_number AS "PAN", gstin_number AS "GSTIN",
                       tan_number AS "TAN", cust_status AS "Status",
                       contact AS "Contact"
                FROM cust_master {where} ORDER BY cust_name ASC''',
            params
        )
        rows = cursor.fetchall()
        _close(cursor, conn)

        buf = io.StringIO()
        if rows:
            writer = csv.DictWriter(buf, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)

        byte_buf  = io.BytesIO(buf.getvalue().encode('utf-8-sig'))
        filename  = f'customers_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        return send_file(byte_buf, mimetype='text/csv',
                         as_attachment=True, download_name=filename)
    except RuntimeError as e:
        return _error(str(e), 503)
    except MySQLError as e:
        return _error(f'Database error: {e}', 500)



# ─────────────────────────────────────────────────────────────
# BOOKS OF ACCOUNTS API ENDPOINTS
# ─────────────────────────────────────────────────────────────

@app.route('/api/books/widgets', methods=['GET'])
def books_widgets():
    # Powers the top five metric summary cards on Books of Accounts screen
    fy = request.args.get('fin_year', '').strip()
    if not fy:
        return jsonify({'error': 'fin_year is required'}), 400
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM vw_books_widgets WHERE fin_year = %s", (fy,))
        row = cursor.fetchone()
        _close(cursor, conn)
        
        if not row:
            return jsonify({'sales_income': 0, 'adjustments': 0, 'linked_adjustments': 0, 'total_income': 0, 'tds_deductible': 0})
        return jsonify(row)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/books/summary', methods=['GET'])
def books_summary():
    # Powers the aggregated "Summary" tab grid view by customer profile safely
    fy = request.args.get('fin_year', '').strip()
    if not fy:
        return jsonify({'error': 'fin_year is required'}), 400
    status_filter = request.args.get('status', '').strip()
        
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        
        query = """
    SELECT
        c.cust_id,
        c.cust_name,
        c.cust_status,
        COALESCE(SUM(i.taxable_amt), 0.00) AS taxable_amt,
        COALESCE(SUM(i.gross_amt), 0.00) AS gross_amt,
        COALESCE(SUM(i.tds_deductible), 0.00) AS tds_deductible,
        COALESCE(SUM(GREATEST(i.gross_amt - i.tds_deductible, 0)), 0.00) AS collection_amt,
        COALESCE(SUM(i.tds_deductible), 0.00) AS tds_receivable
    FROM cust_master c
    LEFT JOIN income_records i
        ON c.cust_id = i.cust_id
        AND i.fin_year = %s
    WHERE 1=1
    """

        params = [fy]
        
        # DYNAMIC STEP: Filter by Active / Inactive status strictly if selected
        if status_filter and status_filter.upper() != 'ALL':
            query += " AND c.cust_status = %s"
            params.append(status_filter.lower())
            
        query += """
            GROUP BY c.cust_id, c.cust_name, c.cust_status
            HAVING (COALESCE(SUM(i.taxable_amt), 0) > 0 OR COALESCE(SUM(i.gross_amt), 0) > 0 OR COALESCE(SUM(i.tax_amt), 0) > 0)
            ORDER BY c.cust_name ASC
        """
        
        page     = max(1, int(request.args.get('page', 1)))
        per_page = max(1, min(500, int(request.args.get('per_page', 500))))
        count_q  = f"SELECT COUNT(*) AS total FROM ({query}) AS _bsc"
        cursor.execute(count_q, tuple(params))
        total = cursor.fetchone()['total']
        query += f" LIMIT {per_page} OFFSET {(page-1)*per_page}"
        cursor.execute(query, tuple(params))
        rows = cursor.fetchall()
        _close(cursor, conn)
        return jsonify({'data': rows, 'total': total, 'page': page, 'per_page': per_page})
    except Exception as e:
        return jsonify({'error': str(e)}), 500





@app.route('/api/books/income-records', methods=['GET'])
def books_income_records():
    fy = request.args.get('fin_year', '').strip()
    if not fy:
        return jsonify({'error': 'fin_year is required'}), 400
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        # DATE_FORMAT forces consistency across both operational ledger arrays
        page     = max(1, int(request.args.get('page', 1)))
        per_page = max(1, min(500, int(request.args.get('per_page', 100))))
        search_q = request.args.get('search', '').strip()
        base_q = """
            SELECT i.income_id, i.upload_ref, cm.cust_code, i.cust_name, i.doc_number, 
                   DATE_FORMAT(i.doc_date, '%Y-%m-%d') AS doc_date, 
                   i.doc_type, i.taxable_amt, i.tax_amt, i.gross_amt, i.entry_source 
            FROM income_records i
            LEFT JOIN cust_master cm ON cm.cust_id = i.cust_id
            WHERE i.fin_year = %s"""
        ir_params = [fy]
        if search_q:
            base_q += " AND (i.cust_name LIKE %s OR i.doc_number LIKE %s)"
            ir_params.extend([f"%{search_q}%", f"%{search_q}%"])
        cursor.execute(f"SELECT COUNT(*) AS total FROM ({base_q}) AS _c", tuple(ir_params))
        total = cursor.fetchone()['total']
        base_q += f" ORDER BY i.cust_name ASC, i.doc_date DESC LIMIT {per_page} OFFSET {(page-1)*per_page}"
        cursor.execute(base_q, tuple(ir_params))
        rows = cursor.fetchall()
        _close(cursor, conn)
        return jsonify({'data': rows, 'total': total, 'page': page, 'per_page': per_page})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
@app.route('/api/books/detailed', methods=['GET'])
def get_books_detailed_real():
    fy = request.args.get('fin_year', '').strip()
    if not fy:
        return jsonify({'error': 'fin_year is required'}), 400
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        # DATE_FORMAT(doc_date, '%Y-%m-%d') forces clean text outputs like '2025-03-02'
        # DATE_FORMAT(doc_date, '%Y-%m-%d') forces clean text outputs like '2025-03-02'
        cursor.execute("""
    SELECT 
        cust_name, 
        doc_number AS invoice_no, 
        DATE_FORMAT(doc_date, '%Y-%m-%d') AS invoice_date, 
        taxable_amt, 
        gross_amt, 
        tds_section, 
        tds_rate, 
       
        tds_deductible, 
        
        (gross_amt - tds_deductible) AS collection_amt, 
        tds_deductible AS tds_receivable, 
        'Automated Sync' AS remarks
    FROM income_records 
    WHERE fin_year = %s AND doc_type = 'invoice'
    ORDER BY doc_date DESC
    """, (fy,))

        rows = cursor.fetchall()
        _close(cursor, conn)
        return jsonify({'data': rows, 'total': len(rows)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500



@app.route('/api/books/unreferenced', methods=['GET'])
def books_unreferenced():
    # Feeds rows to the isolated sorting dashboard ("Unreferenced Entries")
    fy = request.args.get('fin_year', '').strip()
    if not fy:
        return jsonify({'error': 'fin_year is required'}), 400
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT unref_id, upload_ref, raw_identifier, doc_number, doc_date, entry_amt, unref_reason 
            FROM unreferenced_entries 
            WHERE fin_year = %s AND resolved = 0
        """, (fy,))
        rows = cursor.fetchall()
        _close(cursor, conn)
        return jsonify({'data': rows})
    except Exception as e:
        return jsonify({'error': str(e)}), 500




# ─────────────────────────────────────────────────────────────
# TDS & TCS LIVE DATA ROUTING
# ─────────────────────────────────────────────────────────────

@app.route('/api/tds-codes', methods=['GET'])
def get_tds_codes():
    entity_classification = request.args.get('entity', '').strip()
    fin_year = request.args.get('fin_year', '').strip()
    quarter = request.args.get('quarter', '').strip()
    
    # Select fields cleanly without restrictive limiting thresholds
    query = """
        SELECT 
            id, section_code, sub_code, tds_rate, surcharge_pct, 
            cess_pct, appln_form, payee_type, description, effective_from
        FROM tds_codes_master 
        WHERE is_active = 1
    """
    params = []
    
    # Filter dropdown updates from your updated element ID
    # v2 tds_codes_master has no fin_year or quarter columns — filter by payee_type only
    if entity_classification and entity_classification.upper() != 'ALL':
        query += " AND UPPER(TRIM(payee_type)) = %s"
        params.append(entity_classification.upper())
        
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(query, tuple(params))
        rows = cursor.fetchall()
        _close(cursor, conn)
        
        clean_rows = []
        for r in rows:
            date_str = r['effective_from'].strftime('%Y-%m-%d') if r['effective_from'] else '—'
            
            # EXACT FIELD STRUCTURAL DICTIONARY FOR RENDER_TDS_TABLE LOOP MAPPING
            clean_rows.append({
                'gl_code': r['id'],                          # Triggers t.gl_code
                'tds_section': r['section_code'],            # Triggers t.tds_section
                'sub_code': r['sub_code'] or '—',            # Triggers t.sub_code
                'tds_rate': float(r['tds_rate']),            # Triggers t.tds_rate
                'surcharge_pct': float(r['surcharge_pct']),  # Triggers t.surcharge_pct
                'cess_pct': float(r['cess_pct']),            # Triggers t.cess_pct
                'appln_form': r['appln_form'] or '—',        # Triggers t.appln_form
                'applicable_to': r['payee_type'] or 'ALL',   # Triggers t.applicable_to
                'tds_description': r['description'],          # Triggers t.tds_description
                'eff_from_date': date_str                    # Triggers t.eff_from_date
            })
            
        return jsonify({'data': clean_rows, 'total': len(clean_rows)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500



@app.route('/api/tcs-codes', methods=['GET'])
def get_tcs_codes():
    entity_classification = request.args.get('entity', '').strip()
    fin_year = request.args.get('fin_year', '').strip()
    quarter = request.args.get('quarter', '').strip()
    
    query = "SELECT * FROM tcs_codes_master WHERE 1=1"
    params = []
    
    if entity_classification and entity_classification.upper() != 'ALL':
        query += " AND TRIM(applicable_to) = %s"
        params.append(entity_classification)
    if fin_year:
        query += " AND fin_year = %s"
        params.append(fin_year)
    if quarter and quarter.upper() != 'ALL':
        query += " AND quarter = %s"
        params.append(quarter)
        
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(query, tuple(params))
        rows = cursor.fetchall()
        _close(cursor, conn)
        return jsonify({'data': rows, 'total': len(rows)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

   

@app.route('/api/customers/<int:cust_id>/tds', methods=['GET'])
def customer_tds_real(cust_id):
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        
        query = """
            SELECT DISTINCT
                t.id, t.section_code, t.sub_code, t.tds_rate, t.surcharge_pct, 
                t.cess_pct, t.appln_form, t.payee_type, t.description, t.effective_from
            FROM tds_codes_master t
            JOIN income_records ir ON UPPER(TRIM(ir.tds_section)) = UPPER(TRIM(t.section_code))
            WHERE ir.cust_id = %s AND t.is_active = 1
            ORDER BY t.section_code ASC, t.payee_type ASC
        """
        
        cursor.execute(query, (cust_id,))
        rows = cursor.fetchall()
        _close(cursor, conn)
        
        clean_rows = []
        for r in rows:
            date_str = r['effective_from'].strftime('%Y-%m-%d') if r['effective_from'] else '—'
            clean_rows.append({
                'gl_code': r['id'],
                'tds_section': r['section_code'],
                'sub_code': r['sub_code'] or '—',
                'tds_rate': float(r['tds_rate']),
                'surcharge_pct': float(r['surcharge_pct']),
                'cess_pct': float(r['cess_pct']),
                'appln_form': r['appln_form'] or '—',
                'applicable_to': r['payee_type'] or 'ALL',
                'tds_description': r['description'],
                'eff_from_date': date_str
            })
            
        return jsonify({'data': clean_rows, '_status': 'success'}), 200
    except Exception as e:
        return jsonify({'error': str(e), '_status': 'error'}), 500




# ══════════════════════════════════════════════════════════════
#  API — UPLOADS 
# ══════════════════════════════════════════════════════════════

# ═════════════════════════════════════════════════════════════
#  PART 1: PAYLOAD VALIDATION & CRYPTOGRAPHIC DEDUPLICATION
# ═════════════════════════════════════════════════════════════

@app.route('/api/uploads', methods=['POST'])
def process_sales_register_upload():
    if 'file' not in request.files:
        return jsonify({'error': 'No file element detected in payload'}), 400
    file = request.files['file']
    module_target = 'form26as'
    if file.filename == '':
        return jsonify({'error': 'No selected file source detected'}), 400
    file_ext = file.filename.split('.')[-1].lower()
    if file_ext not in ['xlsx', 'xlsm', 'txt', 'pdf']:
        return jsonify({'error': 'Unsupported format. Use a valid .xlsx, .txt, or .pdf file.'}), 400
    try:
        file_bytes = file.read()
        sha256_hash = hashlib.sha256(file_bytes).hexdigest()
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        # For Form 26AS: check for exact hash match only (same file byte-for-byte)
        # We still allow re-processing if needed by checking rows individually
        cursor.execute("SELECT file_name FROM file_uploads_log WHERE file_hash = %s LIMIT 1", (sha256_hash,))
        duplicate_check = cursor.fetchone()
        if duplicate_check:
            _close(cursor, conn)
            return jsonify({'error': f"This file has already been uploaded (identical content detected). Upload a new file to add more records."}), 400
        cursor.close()
        cursor = conn.cursor()
        upload_batch_ref = f"UPL-{uuid.uuid4().hex[:8].upper()}"
        parsed_rows = []

        # =══════════════════════════════════════════════════════════
        # ENGINE A: EXCEL (.XLSX) BINARY SPREADSHEET LOOP
        # =══════════════════════════════════════════════════════════
        if file_ext in ['xlsx', 'xlsm']:
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            sheet = wb.active
            headers = [str(cell.value).strip().lower() for cell in sheet]
            for row_idx in range(2, sheet.max_row + 1):
                row_values = [cell.value for cell in sheet[row_idx]]
                if not any(row_values): 
                    continue
                row_dict = dict(zip(headers, row_values))
                d_raw = row_dict.get('document_date')
                d_str = d_raw.strftime('%Y-%m-%d') if isinstance(d_raw, datetime) else str(d_raw).strip()
                extracted_sec = str(row_dict.get('tds_section') or row_dict.get('section_code') or '').strip().upper()
                tax_category = 'TCS' if extracted_sec.startswith('206') else 'TDS'
                parsed_rows.append({
                    'customer_name': str(row_dict.get('customer_name', '')).strip(),
                    'deductor_name': str(row_dict.get('customer_name', '')).strip(),
                    'deductor_tan': '—',
                    'section_code': extracted_sec,
                    'tax_category': tax_category,
                    'document_number': str(row_dict.get('document_number', '')).strip(),
                    'document_date': d_str,
                    'taxable_amount': float(row_dict.get('taxable_amount') or 0),
                    'tds_deductible': float(row_dict.get('tds_deductible') or 0)
                })

        # =══════════════════════════════════════════════════════════
        # ENGINE B: PLAIN TEXT (.TXT) TAB/COMMA SPLITTING MATRIX
        # =══════════════════════════════════════════════════════════
        elif file_ext == 'txt':
            text_content = file_bytes.decode('utf-8')
            lines = text_content.splitlines()
            for line in lines:
                if not line.strip() or 'customer_name' in line.lower(): 
                    continue
                tokens = re.split(r'\t|,| {2,}', line.strip())
                if len(tokens) >= 6:
                    extracted_sec = tokens[6].strip().upper() if len(tokens) > 6 else ''
                    tax_category = 'TCS' if extracted_sec.startswith('206') else 'TDS'
                    parsed_rows.append({
                        'customer_name': tokens[0].strip(),
                        'deductor_name': tokens[0].strip(),
                        'deductor_tan': '—',
                        'section_code': extracted_sec,
                        'tax_category': tax_category,
                        'document_number': tokens[1].strip(),
                        'document_date': tokens[2].strip(),
                        'taxable_amount': float(tokens[4] or 0),
                        'tds_deductible': float(tokens[5] or 0)
                    })

        # =══════════════════════════════════════════════════════════
        # ENGINE C: PYPDF SMART ADAPTIVE DUAL-LAYOUT SCANNER
        # =══════════════════════════════════════════════════════════
        elif file_ext == 'pdf':
            pdf_reader = PdfReader(io.BytesIO(file_bytes))
            first_page_text = pdf_reader.pages[0].extract_text() or ""
            normalized_first_page = re.sub(r' +', ' ', first_page_text)
            if "annual tax statement" in normalized_first_page.lower() or "traces" in normalized_first_page.lower():
                assessee_match = re.search(r'Name of Assessee\s+([A-Za-z0-9 ]+)', normalized_first_page)
                current_customer_name = assessee_match.group(1).strip() if assessee_match else "Unknown Customer"
                current_deductor_name = "Unknown Deductor"
                current_deductor_tan = "—"
                for page in pdf_reader.pages:
                    page_text = page.extract_text()
                    if not page_text: 
                        continue
                    normalized_text = re.sub(r' +', ' ', page_text)
                    lines = normalized_text.splitlines()
                    for line in lines:
                        line_str = line.strip()
                        if not line_str: 
                            continue
                        # pypdf sometimes splits TAN mid-token e.g. 'PWYC7918 3E'
                        # Collapse only the split digits: e.g. 'PWYC7918 3E' -> 'PWYC79183E'
                        line_nospace = re.sub(r'([A-Z]{4}[0-9]{1,4})\s+([0-9]{1,4}[A-Z])\b', lambda m: m.group(1)+m.group(2), line_str)
                        tan_match = re.search(r'\b([A-Z]{4}[0-9]{5}[A-Z])\b', line_nospace)
                        if tan_match:
                            current_deductor_tan = tan_match.group(1).upper().strip()
                            clean_name = line_nospace.split(tan_match.group(0))[0]
                            for phrase in ["Name of Deductor", "Deductor Name", "TAN of Deductor", ":", "-", "—"]:
                                clean_name = clean_name.replace(phrase, "")
                            current_deductor_name = clean_name.strip() if clean_name.strip() else "Unknown Deductor"
                            continue
                                                
                        section_regex_match = re.search(r'\b(192|194[A-Za-z]?|193|195|206[A-Za-z]?)\b', line_str)
                        
                        if section_regex_match and re.search(r'\d{2}-[A-Za-z]{3}-\d{4}', line_str):
                            # Automatically extracts the exact section found by the regex pattern directly
                            active_section = section_regex_match.group(1).upper().strip()
                            
                            raw_tokens = [t.strip() for t in line_str.split(' ') if t.strip()]
                            tokens = [t for t in raw_tokens if t not in ['-', '—', '/']]
                            dates_found = [t for t in tokens if re.match(r'^\d{2}-[A-Za-z]{3}-\d{4}$', t)]
                            
                            tax_category = 'TCS' if active_section.startswith('206') else 'TDS'
                            
                            numbers_found = []
                            for t in tokens:
                                clean_num = t.replace(',', '')
                                if re.match(r'^\d+\.\d{2}$', clean_num):
                                    numbers_found.append(float(clean_num))

                            if len(dates_found) >= 1 and len(numbers_found) >= 2:
                                try:
                                    date_obj = datetime.strptime(dates_found[0], '%d-%b-%Y')
                                    clean_date_str = date_obj.strftime('%Y-%m-%d')
                                except:
                                    clean_date_str = None
                                parsed_rows.append({
                                    'customer_name': current_customer_name,
                                    'deductor_name': current_deductor_name,
                                    'deductor_tan': current_deductor_tan,
                                    'section_code': active_section,
                                    'tax_category': tax_category,
                                    'document_number': f"TRACES-{upload_batch_ref}",
                                    'document_date': clean_date_str,
                                    'taxable_amount': float(numbers_found[0]),
                                    'tds_deductible': float(numbers_found[-1])
                                })
            else:
                for page in pdf_reader.pages:
                    page_text = page.extract_text()
                    if not page_text: 
                        continue
                    lines = page_text.splitlines()
                    for line in lines:
                        line_str = line.strip()
                        if not line_str or 'customer_name' in line_str.lower() or 'inv-' not in line_str.lower():
                            continue
                        tokens = re.split(r'\t|,|\s{2,}', line_str)
                        tokens = [t.strip() for t in tokens if t.strip()]
                        if len(tokens) >= 6:
                            try:
                                parsed_rows.append({
                                    'customer_name': tokens[0],
                                    'deductor_name': tokens[0],
                                    'deductor_tan': '—',
                                    'section_code': '',
                                    'tax_category': 'TDS',
                                    'document_number': tokens[1],
                                    'document_date': tokens[2].replace('/', '-'),
                                    'taxable_amount': float(tokens[4]),
                                    'tds_deductible': float(tokens[5])
                                })
                            except Exception as parse_err:
                                print(f"Skipping row anomalies: {parse_err}")

        # ══════════════════════════════════════════════════════════
        # PART 2: DYNAMIC AUTOMATED FY DETECTOR & TRANSACTIONS PIPELINE
        # ══════════════════════════════════════════════════════════
        if not parsed_rows:
            return jsonify({'error': 'No data fields could be extracted from this document.'}), 400
            
             
        
        try:
            # Added [0] index to cleanly target the first parsed data row safely
            sample_date_raw = parsed_rows[0].get('document_date')
            if not sample_date_raw:
                raise ValueError("First row document_date is empty.")
                
            sample_date_parts = str(sample_date_raw).split('-')
            sample_year = int(sample_date_parts[0])
            sample_month = int(sample_date_parts[1])
            
            # Written with names to completely bypass the chat tag-stripper bug
            months_list_array = list((1, 2, 3))
            
            if sample_month in months_list_array:
                fin_year = f"{sample_year - 1}-{str(sample_year)[2:]}"
            else:
                fin_year = f"{sample_year}-{str(sample_year + 1)[2:]}"
            # Map month to financial quarter using a dictionary lookup
            quarter_map = {4: "Q1", 5: "Q1", 6: "Q1", 7: "Q2", 8: "Q2", 9: "Q2", 10: "Q3", 11: "Q3", 12: "Q3", 1: "Q4", 2: "Q4", 3: "Q4"}
            calculated_quarter = quarter_map.get(sample_month, "Q1")

        except Exception as fy_err:
            _close(cursor, conn)
            return jsonify({'error': f'Could not determine Financial Year from file dates. Detail: {str(fy_err)}'}), 400



        # ══════════════════════════════════════════════════════════
        # PART 3: DYNAMIC RECON SILENT RE-MAPPING INJECTION LAYER
        # ══════════════════════════════════════════════════════════
        inserted_rows_count = 0
        tan_map = {}
        try:
            cursor.execute("SELECT LOWER(TRIM(cust_name)) AS name, tan_number FROM cust_master")
            for cm_row in cursor.fetchall():
                if cm_row['name'] and cm_row['tan_number']:
                    tan_map[cm_row['name']] = cm_row['tan_number']
        except Exception as map_err:
            print(f"Customer map extraction bypassed: {map_err}")

        seen_records = set()

        for item in parsed_rows:
            doc_date_raw = item.get('document_date')
            resolved_tan = item.get('deductor_tan', '—')
            cust_name_str = str(item.get('customer_name', '')).lower().strip()
            deduct_name_str = str(item.get('deductor_name', '')).lower().strip()
            sec_code = item.get('section_code') or None
            tax_amt = float(item.get('taxable_amount', 0.00))
            tds_ded = float(item.get('tds_deductible', 0.00))

            if resolved_tan == '—' or not resolved_tan or deduct_name_str == 'unknown deductor':
                resolved_tan = tan_map.get(cust_name_str, tan_map.get(deduct_name_str, '—'))
                
            final_deductor_name = item.get('deductor_name', 'Unknown Deductor')
            if final_deductor_name == 'Unknown Deductor' or not final_deductor_name:
                final_deductor_name = item.get('customer_name', 'Unknown Deductor')

            record_fingerprint = f"{resolved_tan}_{doc_date_raw}_{sec_code}_{tax_amt}_{tds_ded}"
            if record_fingerprint in seen_records:
                continue
            seen_records.add(record_fingerprint)

            # Check if this exact record already exists
            cursor.execute(
                "SELECT record_id FROM form26as_records WHERE tan_of_deductor=%s AND section_code=%s AND transaction_date=%s AND tds_credited=%s AND fin_year=%s LIMIT 1",
                (resolved_tan, sec_code, doc_date_raw, tds_ded, fin_year)
            )
            if cursor.fetchone():
                continue  # skip exact duplicate
            cursor.execute(
                "INSERT INTO form26as_records (fin_year,tan_of_deductor,deductor_name,section_code,transaction_date,booking_amt,tds_credited) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                (
                fin_year, resolved_tan, final_deductor_name,
                sec_code, doc_date_raw, tax_amt, tds_ded))
            inserted_rows_count += 1

        file_size_kb = round(len(file_bytes) / 1024.0, 2)
        cursor.execute("""
            INSERT INTO file_uploads_log (file_name, file_hash, fin_year, quarter, file_size_kb, upload_status, rows_imported, module_name)
            VALUES (%s, %s, %s, %s, %s, 'success', %s, %s)
        """, (file.filename, sha256_hash, fin_year, calculated_quarter or 'Q1', file_size_kb, inserted_rows_count, module_target)) # type: ignore
        cursor.execute(
        'INSERT INTO audit_logs (action_type, module_name, description, severity_level, ip_address) '
        'VALUES ("UPLOAD", "FORM26AS", %s, "INFO", %s)',
            (
        f"Form 26AS upload FY {fin_year}: {inserted_rows_count} inserted, "
        f"{len(parsed_rows) - inserted_rows_count} skipped.",
        request.remote_addr or 'unknown'
            )
        )
        conn.commit()
        _close(cursor, conn)
        # Update tds_deductible using section rates, then sync to BigQuery
        update_tds_deductible_from_section(fin_year)
        bq_sync_form26as(fin_year)
        bq_sync_tds_summary(fin_year)
        return jsonify({
            'status': 'success',
            'message': 'File parsed, mapped, deduplicated, and database records synchronized successfully.',
            'parsed_count': inserted_rows_count,
            'upload_ref': upload_batch_ref
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Upload parsing failed: {str(e)}'}), 500

       


# ─────────────────────────────────────────────────────────────────────────────────
#  CUSTOMER MASTER — EXCEL UPLOAD
#  Columns: Customer Code | Customer Name | PAN | GSTIN | TAN | Status | Contact
# ─────────────────────────────────────────────────────────────────────────────────        

@app.route('/api/customers/upload-excel', methods=['POST'])
def upload_customer_excel():
    if 'file' not in request.files:
        return jsonify({'error': 'No file in request'}), 400
    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'No file selected'}), 400
    ext = file.filename.rsplit('.', 1)[-1].lower()
    if ext not in ('xlsx', 'xlsm'):
        return jsonify({'error': 'Customer Master upload requires an .xlsx file'}), 400
    try:
        file_bytes = file_bytes_raw = file.read()
        sha256 = hashlib.sha256(file_bytes).hexdigest()
        conn = get_db()
        cur  = conn.cursor(dictionary=True)
        cur.execute('SELECT upload_id FROM file_uploads_log WHERE file_hash = %s LIMIT 1', (sha256,))
        if cur.fetchone():
            _close(cur, conn)
            return jsonify({'error': 'This file has already been uploaded.'}), 400

        wb    = load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheet = wb.active
        raw_h = [str(v).strip().lower().replace(' ','_') if v is not None else ''
                 for v in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]

        def _col(keys):
            for k in keys:
                for h in raw_h:
                    if k in h: return h
            return None

        col_code   = _col(['code'])
        col_name   = _col(['name'])
        col_pan    = _col(['pan'])
        col_gstin  = _col(['gstin','gst'])
        col_tan    = _col(['tan'])
        col_status = _col(['status'])
        col_contact= _col(['contact','phone','mobile'])

        PAN_RE = re.compile(r'^[A-Z]{5}[0-9]{4}[A-Z]$')
        TAN_RE = re.compile(r'^[A-Z]{4}[0-9]{5}[A-Z]$')

        inserted = updated = 0
        errors = []
        cur.close()
        cur = conn.cursor()

        for rn, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): continue
            # values_only=True gives plain Python values (str/int/float/None), not Cell objects
            rd = dict(zip(raw_h, [v for v in row]))

            cust_code    = str(rd.get(col_code)    or '').strip()
            cust_name    = str(rd.get(col_name)    or '').strip()
            pan_number   = str(rd.get(col_pan)     or '').strip().upper()
            gstin_number = str(rd.get(col_gstin)   or '').strip().upper() or None
            tan_number   = str(rd.get(col_tan)     or '').strip().upper()
            cust_status  = str(rd.get(col_status)  or 'active').strip().lower()
            contact      = str(rd.get(col_contact) or '').strip() or None

            if not cust_name or not pan_number or not tan_number:
                errors.append(f'Row {rn}: Missing Name, PAN or TAN — skipped')
                continue
            if not PAN_RE.match(pan_number):
                errors.append(f'Row {rn}: Invalid PAN {pan_number} — skipped')
                continue
            if not TAN_RE.match(tan_number):
                errors.append(f'Row {rn}: Invalid TAN {tan_number} — skipped')
                continue
            if cust_status not in ('active','inactive'):
                cust_status = 'active'
            if not cust_code:
                cust_code = f'CUST{rn:04d}'

            cur.execute('SELECT cust_id FROM cust_master WHERE pan_number = %s', (pan_number,))
            existing = cur.fetchone()
            if existing:
                cur.execute('SELECT cust_id FROM cust_master WHERE tan_number = %s AND pan_number != %s', (tan_number, pan_number))
                if cur.fetchone():
                    errors.append(f'Row {rn}: TAN {tan_number} belongs to another customer — skipped')
                    continue
                cur.execute('UPDATE cust_master SET cust_name=%s,tan_number=%s,gstin_number=%s,cust_status=%s,contact=%s WHERE pan_number=%s',
                            (cust_name, tan_number, gstin_number, cust_status, contact, pan_number))
                updated += 1
            else:
                cur.execute('SELECT cust_id FROM cust_master WHERE tan_number = %s', (tan_number,))
                if cur.fetchone():
                    errors.append(f'Row {rn}: TAN {tan_number} already exists — skipped')
                    continue
                cur.execute('SELECT cust_id FROM cust_master WHERE cust_code = %s', (cust_code,))
                if cur.fetchone():
                    cust_code = f'{cust_code}-{rn}'
                cur.execute('INSERT INTO cust_master (cust_code,cust_name,pan_number,gstin_number,tan_number,cust_status,contact) VALUES (%s,%s,%s,%s,%s,%s,%s)',
                            (cust_code, cust_name, pan_number, gstin_number, tan_number, cust_status, contact))
                inserted += 1

                
        file_size_kb = round(len(file_bytes)/1024.0, 2)
        upload_ref = f'CUST-{uuid.uuid4().hex[:8].upper()}'
        
        cur.execute("""
            INSERT INTO file_uploads_log 
            (file_name, file_hash, fin_year, quarter, file_size_kb, upload_status, rows_imported, module_name) 
            VALUES (%s, %s, 'N/A', 'NA', %s, %s, %s, 'customer_master')
        """, (
            file.filename,   
            sha256,          
            file_size_kb,    
            "success",       
            inserted+updated 
        ))


        cur.execute('INSERT INTO audit_logs (action_type,module_name,description,severity_level,ip_address) VALUES ("UPLOAD","CUSTOMER_MASTER",%s,"INFO",%s)',
                    (f"Customer upload: {inserted} inserted, {updated} updated, {len(errors)} skipped.", request.remote_addr or 'unknown'))
        conn.commit()
        _close(cur, conn)
        # Sync to BigQuery after successful Customer Master upload
        bq_sync_cust_master()
        return jsonify({'status':'success','upload_ref':upload_ref,'inserted':inserted,'updated':updated,'skipped':len(errors),'errors':errors[:10],'message':f'{inserted} customers added, {updated} updated, {len(errors)} skipped.'}), 200
    except Exception as e:
        return jsonify({'error': f'Customer upload failed: {str(e)}'}), 500


@app.route('/api/customers/template', methods=['GET'])
def download_customer_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook(); ws = wb.active; ws.title = 'Customer Master'
    headers = ['Customer Code','Customer Name','PAN Number','GSTIN','TAN Number','Status','Contact']
    sample  = ['CUST001','Example Company Pvt Ltd','ABCDE1234F','27ABCDE1234F1Z5','ABCD12345E','active','9876543210']
    hfill = PatternFill('solid', fgColor='0F1E3D')
    hfont = Font(color='FFFFFF', bold=True)
    for col,(h,s) in enumerate(zip(headers,sample),1):
        hc = ws.cell(row=1,column=col,value=h)
        hc.fill, hc.font, hc.alignment = hfill, hfont, Alignment(horizontal='center')
        ws.cell(row=2,column=col,value=s)
        ws.column_dimensions[ws.cell(row=1,column=col).column_letter].width = max(len(h),len(s))+4
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='customer_master_template.xlsx')



# ─────────────────────────────────────────────────────────────
#  BOOKS OF ACCOUNTS — EXCEL / TXT UPLOAD
#  Excel columns: Customer Name | PAN | TAN | Document No |
#  Document Date | Document Type | Taxable Amount | Tax Amount | Gross Amount
# ─────────────────────────────────────────────────────────────


@app.route('/api/books/upload', methods=['POST'])
def upload_books_data():
    if 'file' not in request.files:
        return jsonify({'error': 'No file in request'}), 400
    file     = request.files['file']
    fin_year = request.form.get('fin_year', '').strip()
    if not file.filename:
        return jsonify({'error': 'No file selected'}), 400
    ext = file.filename.rsplit('.', 1)[-1].lower()
    if ext not in ('xlsx', 'xlsm', 'txt'):
        return jsonify({'error': 'Books upload requires .xlsx or .txt file'}), 400

    def _date(raw):
        if not raw or str(raw).strip() in ('', 'None'): return None
        raw = str(raw).strip()
        from datetime import datetime as _dt
        for fmt in ('%Y-%m-%d','%d-%m-%Y','%d/%m/%Y','%m/%d/%Y','%d-%b-%Y','%Y/%m/%d'):
            try: return _dt.strptime(raw, fmt).strftime('%Y-%m-%d')
            except: pass
        return None

    def _detect_fy(date_str):
        if not date_str: return None
        try:
            yr, mo = int(str(date_str)[:4]), int(str(date_str)[5:7])
            return f"{yr}-{str(yr+1)[2:]}" if mo >= 4 else f"{yr-1}-{str(yr)[2:]}"
        except: return None

    try:
        file_bytes = file.read()
        sha256 = hashlib.sha256(file_bytes).hexdigest()
        conn = get_db(); cur = conn.cursor(dictionary=True)
        cur.execute('SELECT upload_id FROM file_uploads_log WHERE file_hash = %s LIMIT 1', (sha256,))
        if cur.fetchone():
            _close(cur, conn)
            return jsonify({'error': 'This file has already been uploaded.'}), 400

        # Load customer lookups
        cur.execute('SELECT cust_id, cust_code, cust_name, pan_number, tan_number FROM cust_master')
        all_custs = cur.fetchall()
        code_map = {c['cust_code'].strip().upper(): c for c in all_custs}
        tan_map  = {c['tan_number'].upper(): c for c in all_custs}
        pan_map  = {c['pan_number'].upper(): c for c in all_custs}
        name_map = {c['cust_name'].strip().lower(): c for c in all_custs}

        # Load TDS rates keyed by section_code (v2 schema — no cust_id column)
        cur.execute("""
            SELECT section_code, tds_rate
            FROM tds_codes_master
            WHERE is_resident = 1 AND is_active = 1
            ORDER BY
                section_code,
                CASE payee_type
                    WHEN 'OTHERS'         THEN 1
                    WHEN 'ALL'            THEN 2
                    WHEN 'INDIVIDUAL_HUF' THEN 3
                    ELSE 4
                END,
                effective_from DESC
        """)
        tds_by_section = {}
        for t in cur.fetchall():
            sc = (t['section_code'] or '').strip().upper()
            if sc and sc not in tds_by_section:
                tds_by_section[sc] = float(t['tds_rate'])

        upload_ref = f'BOOKS-{uuid.uuid4().hex[:8].upper()}'
        rows_data  = []

        if ext in ('xlsx', 'xlsm'):
            wb    = load_workbook(io.BytesIO(file_bytes), data_only=True)
            sheet = wb.active
            headers = [str(v).strip().lower().replace(' ','_') if v is not None else f'col{i}'
                       for i,v in enumerate(next(sheet.iter_rows(min_row=1,max_row=1,values_only=True)))]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(v for v in row if v is not None): continue
                rows_data.append(dict(zip(headers, [str(v).strip() if v is not None else '' for v in row])))
        elif ext == 'txt':
            import csv as _csv
            text = file_bytes.decode('utf-8-sig')
            dialect = 'excel-tab' if chr(9) in text.split('\n')[0] else 'excel'
            reader = _csv.DictReader(io.StringIO(text), dialect=dialect)
            for row in reader:
                clean = {}
                for k, v in row.items():
                    if k is None: continue  # skip extra tab columns
                    key = k.strip().lower().replace(' ','_')
                    if not key: continue
                    clean[key] = str(v).strip() if v is not None else ''
                if any(clean.values()):
                    rows_data.append(clean)

        if not rows_data:
            _close(cur, conn)
            return jsonify({'error': 'No data rows found in the file'}), 400

        # Auto-detect FY from document dates if not provided
        if not fin_year:
            for row in rows_data[:10]:  # check first 10 rows
                raw_d = ''
                for h in row:
                    if any(k in h for k in ['date','doc_date','document_date','invoice_date']):
                        raw_d = str(row[h] or '').strip()
                        break
                detected = _detect_fy(_date(raw_d) if raw_d else None)
                if detected:
                    fin_year = detected
                    break
            if not fin_year:
                _close(cur, conn)
                return jsonify({'error': 'Could not detect Financial Year from document dates. Please check your file.'}), 400

        def _get(row, *keys, default=''):
            # Exact match first — prevents 'tax' colliding with 'taxable_amount'
            for k in keys:
                if k in row:
                    v = row[k]
                    return str(v).strip() if v is not None else default
            # Substring fallback for loose aliases
            for k in keys:
                for h in row:
                    if h and k in h:
                        v = row[h]
                        return str(v).strip() if v is not None else default
            return default

        def _amt(row, *keys):
            v = _get(row, *keys, default='0')
            try:
                num = str(v).replace(',','').replace('₹','').strip()
                return float(num) if num else 0.0
            except: return 0.0

        cur.close()
        cur = conn.cursor(dictionary=True)
        inserted_books = 0; unref_count = 0; row_errors = []

        for i, row in enumerate(rows_data, start=2):
            cust_code_raw = _get(row, 'customer_code','cust_code','code').upper()
            cust_name_raw = _get(row, 'customer_name','cust_name','name')
            pan_raw = _get(row, 'pan').upper()
            tan_raw = _get(row, 'tan').upper()
            doc_number = _get(row, 'document_no','doc_number','invoice_no','doc_no')
            doc_date = _date(_get(row, 'document_date','doc_date','date','invoice_date'))
            _raw_type = (_get(row, 'document_type','doc_type','type') or '').lower().strip()
            _type_map = {
                'invoice': 'invoice', 'inv': 'invoice', 'sales': 'invoice', 'bill': 'invoice',
                'credit_note': 'credit_note', 'credit note': 'credit_note', 'cn': 'credit_note',
                'debit_note': 'debit_note', 'debit note': 'debit_note', 'dn': 'debit_note',
                'advance': 'advance', 'adv': 'advance',
                'other': 'other', 'others': 'other',
            }
            doc_type = _type_map.get(_raw_type, 'invoice')
            taxable_amt = _amt(row, 'taxable_amount','taxable_amt','taxable')
            tax_amt = _amt(row, 'tax_amount','tax_amt','tax')
            gross_amt = _amt(row, 'gross_amount','gross_amt','gross') or (taxable_amt + tax_amt)

            if taxable_amt == 0 and gross_amt > 0:
                taxable_amt = max(gross_amt - tax_amt, 0)

            if not doc_number or not doc_date:
                row_errors.append(f'Row {i}: Missing Document No or Date — skipped')
                continue

            # Per-row FY — use the actual row date's FY, not the file-level detected FY
            # This allows a single file to contain records across multiple FYs correctly
            row_fy = _detect_fy(doc_date)
            effective_fy = row_fy if row_fy else fin_year

            # Customer Code is the primary match key; fall back to TAN, PAN, Name
            matched = (code_map.get(cust_code_raw) or tan_map.get(tan_raw)
                       or pan_map.get(pan_raw) or name_map.get(cust_name_raw.strip().lower()))

            if not matched:
                reason = 'no_tan' if tan_raw else ('no_pan' if pan_raw else 'no_customer')
                cur.execute(
                    'INSERT INTO unreferenced_entries (upload_ref,fin_year,raw_identifier,doc_number,doc_date,entry_amt,unref_reason,remarks) '
                    'VALUES (%s,%s,%s,%s,%s,%s,%s,%s)',
                    (upload_ref, fin_year, cust_name_raw or tan_raw or pan_raw, doc_number, doc_date, gross_amt, reason,
                     f'No match: name={cust_name_raw} PAN={pan_raw} TAN={tan_raw}')
                )
                unref_count += 1
                continue

            cust_id = matched['cust_id']
            
            # ════════════════════════════════════════════════════════
            # DYNAMIC TDS SECTION PARSING ENGINE
            # ════════════════════════════════════════════════════════
            excel_section = _get(row, 'tds_section', 'section_code', 'section', 'sec').strip().upper()
            excel_section = excel_section if excel_section not in ('—', '-', 'NONE', '') else ''

            if excel_section:
                tds_section = excel_section
                tds_rate    = tds_by_section.get(tds_section, 0.0)
            else:
                # No section in file — update_tds_deductible_from_section() fills this
                # after Form 26AS is uploaded and matched
                tds_section = '—'
                tds_rate    = 0.0

            tds_deductible = round(taxable_amt * tds_rate / 100, 2)

            # Auto-detect entry_source from file type and content — never hardcoded
            explicit_source = (_get(row, 'entry_source', 'source') or '').strip().lower()
            if ext == 'txt':
                # TXT files are always sales register exports (Tally/TRACES format)
                entry_source = 'sales_register'
            elif explicit_source and explicit_source not in ('', 'none', '—', '-'):
                # Explicit column in xlsx takes priority
                entry_source = explicit_source
            else:
                # Auto-detect from headers: sales register files typically have
                # voucher/ledger/narration columns; manual files have doc_number/taxable_amount
                sales_register_indicators = {
                    'voucher_no', 'voucher_number', 'voucher_type',
                    'ledger', 'ledger_name', 'narration',
                    'debit', 'credit', 'debit_amount', 'credit_amount',
                    'tally', 'particulars', 'vch_no'
                }
                row_headers = set(row.keys())
                is_sales_register = bool(row_headers & sales_register_indicators)
                entry_source = 'sales_register' if is_sales_register else 'manual'

            # Delete any existing record for this doc_number + cust_id regardless of FY
            # This ensures no duplicates when FY changes between uploads
            cur.execute(
                "DELETE FROM income_records WHERE doc_number=%s AND cust_id=%s",
                (doc_number, cust_id)
            )
            deleted = cur.rowcount
            # Always insert fresh with correct effective_fy
            cur.execute(
                "INSERT INTO income_records "
                "(upload_ref,fin_year,cust_id,cust_name,doc_number,doc_date,doc_type,"
                "taxable_amt,tax_amt,gross_amt,tds_section,tds_rate,tds_deductible,entry_source) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (upload_ref, effective_fy, cust_id, matched['cust_name'], doc_number, doc_date,
                 doc_type, taxable_amt, tax_amt, gross_amt, tds_section, tds_rate,
                 tds_deductible, entry_source)
            )
            if deleted:
                row_errors.append(f'Row {i}: {doc_number} refreshed in FY {effective_fy} (previous record replaced)')
            else:
                inserted_books += 1

        file_size_kb = round(len(file_bytes)/1024.0, 2)
        cur.execute(
            'INSERT INTO file_uploads_log (file_name,file_hash,fin_year,quarter,file_size_kb,upload_status,rows_imported,module_name) '
            'VALUES (%s,%s,%s,"Q1",%s,"success",%s,"books_of_accounts")',
            (file.filename, sha256, fin_year, file_size_kb, inserted_books+unref_count)
        )
        cur.execute(
            'INSERT INTO audit_logs (action_type,module_name,description,severity_level,ip_address) '
            'VALUES ("UPLOAD","BOOKS_OF_ACCOUNTS",%s,"INFO",%s)',
            (f"Books upload FY {fin_year}: {inserted_books} inserted, {unref_count} unmatched, {len(row_errors)} skipped.",
             request.remote_addr or 'unknown')
        )
        conn.commit()
        _close(cur, conn)
        # Sync to BigQuery after successful Books upload
        bq_sync_income_records(fin_year)
        bq_sync_tds_summary(fin_year)
        return jsonify({'status':'success','upload_ref':upload_ref,'fin_year':fin_year,'inserted':inserted_books,'unmatched':unref_count,'skipped':len(row_errors),'errors':row_errors[:10],'message':f'{inserted_books} income records saved. {unref_count} rows unmatched. {len(row_errors)} skipped.'}), 200
    except Exception as e:
        return jsonify({'error': f'Books upload failed: {str(e)}'}), 500

# BOOKS - TEMPLATE (EXCEL)
@app.route('/api/books/template', methods=['GET'])
def download_books_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook(); ws = wb.active; ws.title = 'Books of Accounts'
    headers = ['Customer Code','Document No','Document Date','Document Type','Taxable Amount','Tax Amount','Gross Amount']
    sample  = ['CUST0001','INV-2024-001','2024-04-15','invoice','100000','18000','118000']
    hfill = PatternFill('solid', fgColor='0F1E3D')
    hfont = Font(color='FFFFFF', bold=True)
    for col,(h,s) in enumerate(zip(headers,sample),1):
        hc = ws.cell(row=1,column=col,value=h)
        hc.fill, hc.font, hc.alignment = hfill, hfont, Alignment(horizontal='center')
        ws.cell(row=2,column=col,value=s)
        ws.column_dimensions[ws.cell(row=1,column=col).column_letter].width = max(len(h),len(s))+4
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='books_of_accounts_template.xlsx')
    
    
# BOOKS - TEMPLATE (TEXT)  
@app.route('/api/books/template-txt', methods=['GET'])
def download_books_txt_template():
    try:
        headers = [
            "Customer Code", "Document No",
            "Document Date", "Document Type", "Taxable Amount",
            "Tax Amount", "Gross Amount"
        ]
        txt_content = "\t".join(headers) + "\n"
        
        from flask import Response
        return Response(
            txt_content,
            mimetype="text/tab-separated-values",
            headers={"Content-Disposition": "attachment;filename=books_template.txt"}
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500




# ─────────────────────────────────────────────────────────────
#  FORM 26AS ONLY — HISTORIC LOG RETRIEVER SYSTEM
# ─────────────────────────────────────────────────────────────


@app.route('/api/uploads', methods=['GET'])
def get_uploads_history_for_frontend():
    module   = request.args.get('module',   '').strip()
    fin_year = request.args.get('fin_year', '').strip()

    clauses, params = [], []
    if module:
        clauses.append('LOWER(module_name) = %s')
        params.append(module.lower())
    if fin_year:
        clauses.append('fin_year = %s')
        params.append(fin_year)
    where = ('WHERE ' + ' AND '.join(clauses)) if clauses else ''

    query = f"""
        SELECT
            LOWER(SUBSTRING_INDEX(file_name, '.', -1)) AS file_ext,
            file_name AS orig_filename,
            module_name AS upload_module,
            fin_year,
            quarter AS quarter_ref,
            file_size_kb,
            upload_status AS parse_status,
            rows_imported AS row_count_total,
            uploaded_by AS uploaded_by_name,
            DATE_FORMAT(uploaded_at, '%Y-%m-%d %H:%i') AS uploaded_at
        FROM file_uploads_log
        {where}
        ORDER BY upload_id DESC
    """
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(query, params)
        rows = cursor.fetchall()
        _close(cursor, conn)
        return jsonify({'data': rows, 'total': len(rows)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
         



# ─────────────────────────────────────────────────────────────
# DYNAMIC PRODUCTION — RECONCILIATION API CONTROLLERS

# LIVE OVERVIEW WIDGET STATISTICS CONTROLLER — WITH PREFIX STRIPPER
# ─────────────────────────────────────────────────────────────

@app.route('/api/recon/stats', methods=['GET'])
def get_recon_stats():
    raw_fy = request.args.get('fin_year', '').strip() or request.args.get('fy', '').strip()
    if not raw_fy:
        return jsonify({'error': 'fin_year is required'}), 400
        
    # Isolate the core numeric sequence (e.g., "2022-23")
    fy_numeric = raw_fy.upper().replace('FY', '').replace(' ', '').strip()
    if not fy_numeric:
        return jsonify({'error': 'Invalid financial year layout'}), 400
        
    fy = fy_numeric
    
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        
        # 1. Calculate cumulative overview total balances across structural tables
        cursor.execute("""
            SELECT 
                COALESCE(SUM(b.system_tds), 0.00) AS total_system_tds,
                COALESCE(SUM(g.f26as_credit), 0.00) AS total_f26as_credit,
                COALESCE(SUM(b.gross_income), 0.00) AS total_book_credit,
                CASE 
                    WHEN COALESCE(SUM(b.system_tds), 0) = 0 THEN 0.00
                    ELSE (COALESCE(SUM(g.f26as_credit), 0.00) / COALESCE(SUM(b.system_tds), 0.00)) * 100.00
                END AS pct_credit
            FROM cust_master c
            LEFT JOIN ( 
                SELECT cust_id, SUM(tds_deductible) AS system_tds, SUM(gross_amt) AS gross_income
                FROM income_records
                WHERE fin_year = %s
                GROUP BY cust_id
            ) b ON c.cust_id = b.cust_id
            LEFT JOIN (
                SELECT LOWER(TRIM(tan_of_deductor)) AS clean_tan, 
                       SUM(tds_credited) AS f26as_credit 
                FROM form26as_records 
                WHERE fin_year = %s
                GROUP BY LOWER(TRIM(tan_of_deductor))
            ) g ON (LOWER(TRIM(c.tan_number)) = g.clean_tan OR LOWER(TRIM(c.pan_number)) = g.clean_tan)
        """, (fy, fy))
        
        totals = cursor.fetchone() or {
            'total_system_tds': 0.00, 'total_f26as_credit': 0.00, 'total_book_credit': 0.00, 'pct_credit': 0.00
        }
        
        # 2. Compute true dynamic category row item counts matching database realities with overrides
        cursor.execute("""
            SELECT recon_status, COUNT(*) AS cnt_val
            FROM (
                SELECT 
                    COALESCE(ro.manual_status, CASE 
                        WHEN COALESCE(b.system_tds, 0) = 0 AND COALESCE(g.f26as_credit, 0) = 0 AND COALESCE(b.gross_income, 0) = 0 THEN 'open'
                        WHEN ABS(COALESCE(g.f26as_credit, 0.00) - COALESCE(b.system_tds, 0.00)) < 5.00 THEN 'reconciled'
                        WHEN COALESCE(g.f26as_credit, 0.00) = 0.00 THEN 'open'
                        WHEN (CASE WHEN COALESCE(b.system_tds, 0) >= COALESCE(g.f26as_credit, 0) THEN (COALESCE(g.f26as_credit, 0) / b.system_tds) * 100 ELSE (b.system_tds / g.f26as_credit) * 100 END) >= 85.00 THEN 'likely_match'
                        WHEN (CASE WHEN COALESCE(b.system_tds, 0) >= COALESCE(g.f26as_credit, 0) THEN (COALESCE(g.f26as_credit, 0) / b.system_tds) * 100 ELSE (b.system_tds / g.f26as_credit) * 100 END) >= 50.00 THEN 'suggested_match'
                        ELSE 'open'
                    END) AS recon_status
                FROM cust_master c
                LEFT JOIN (
                    SELECT cust_id, SUM(tds_deductible) AS system_tds, SUM(gross_amt) AS gross_income
                    FROM income_records
                    WHERE fin_year = %s
                    GROUP BY cust_id
                ) b ON c.cust_id = b.cust_id
                LEFT JOIN (
                    SELECT LOWER(TRIM(tan_of_deductor)) AS clean_tan, SUM(tds_credited) AS f26as_credit 
                    FROM form26as_records 
                    WHERE fin_year = %s
                    GROUP BY LOWER(TRIM(tan_of_deductor))
                ) g ON (LOWER(TRIM(c.tan_number)) = g.clean_tan OR LOWER(TRIM(c.pan_number)) = g.clean_tan)
                LEFT JOIN recon_overrides ro ON ro.cust_id = c.cust_id AND ro.fin_year = %s
                WHERE (b.system_tds IS NOT NULL OR g.f26as_credit IS NOT NULL)
            ) AS final_counts_matrix
            GROUP BY recon_status
        """, (fy, fy, fy))
        
        count_rows = cursor.fetchall()
        status_map = {r['recon_status']: r['cnt_val'] for r in count_rows}
        
        # 3. Calculate dynamic Grand Total Metrics safely
        cursor.execute("""
            SELECT COUNT(DISTINCT c.cust_id) AS total_cust
            FROM cust_master c
            LEFT JOIN (
                SELECT DISTINCT cust_id
                FROM income_records 
                WHERE fin_year = %s
            ) b ON c.cust_id = b.cust_id
            LEFT JOIN (
                SELECT LOWER(TRIM(tan_of_deductor)) AS t 
                FROM form26as_records 
                WHERE fin_year = %s
            ) g ON LOWER(TRIM(c.tan_number)) = g.t
            WHERE b.cust_id IS NOT NULL OR g.t IS NOT NULL
        """, (fy, fy))
        
        grand_total_row = cursor.fetchone()
        grand_total = grand_total_row['total_cust'] if grand_total_row else 0
        
        _close(cursor, conn)
        
        return jsonify({
            'total_system_tds': float(totals['total_system_tds']),
            'total_f26as_credit': float(totals['total_f26as_credit']),
            'total_book_credit': float(totals['total_book_credit']),
            'pct_credit': float(totals['pct_credit']),
            'cnt_grand_total': grand_total,
            'cnt_open': status_map.get('open', 0),
            'cnt_reconciled': status_map.get('reconciled', 0),
            'cnt_likely_match': status_map.get('likely_match', 0),
            'cnt_suggested_match': status_map.get('suggested_match', 0),
            'cnt_reserved': status_map.get('reserved', 0),
            'cnt_not_liable': status_map.get('not_liable', 0),
            'cnt_disabled': status_map.get('disabled', 0)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

          
     
        

# ─────────────────────────────────────────────────────────────
# DYNAMIC RECONCILIATION DATA ENGINE WITH ROW SEARCH FILTERING
# ─────────────────────────────────────────────────────────────
               
      
@app.route('/api/recon/records', methods=['GET'])
def get_recon_records():
    raw_fy = request.args.get('fin_year', '').strip() or request.args.get('fy', '').strip()
    status_filter = request.args.get('status', '').strip()
    search_q = request.args.get('search', '').strip()
    
    if not raw_fy:
        return jsonify({'error': 'fin_year is required'}), 400
        
    fy_numeric = raw_fy.upper().replace('FY', '').replace(' ', '').strip()
    if not fy_numeric:
        return jsonify({'error': 'Invalid financial year layout'}), 400
    fy = fy_numeric
    
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        params = [fy, fy, fy]
        
        query = """
            SELECT * FROM (
                SELECT 
                    c.cust_id AS recon_id, 
                    c.cust_id,
                    c.cust_name, 
                    c.pan_number, 
                    c.tan_number,
                    COALESCE(b.system_tds, 0.00) AS system_tds_amt,
                    COALESCE(g.f26as_credit, 0.00) AS f26as_credit_amt,
                    COALESCE(b.gross_income, 0.00) AS book_credit_amt,
                    (COALESCE(b.system_tds, 0.00) - COALESCE(g.f26as_credit, 0.00)) AS variance_amt,
                    CASE 
                        WHEN COALESCE(b.system_tds, 0) = 0 AND COALESCE(g.f26as_credit, 0) = 0 THEN 100.00
                        WHEN ABS(COALESCE(b.system_tds, 0.00) - COALESCE(g.f26as_credit, 0.00)) < 5.00 THEN 100.00
                        WHEN COALESCE(b.system_tds, 0) = 0 OR COALESCE(g.f26as_credit, 0) = 0 THEN 0.00
                        WHEN b.system_tds >= g.f26as_credit THEN (g.f26as_credit / b.system_tds) * 100
                        ELSE (b.system_tds / g.f26as_credit) * 100
                    END AS match_score,
                    COALESCE(ro.manual_status, CASE 
                        WHEN COALESCE(b.system_tds, 0) = 0 AND COALESCE(g.f26as_credit, 0) = 0 AND COALESCE(b.gross_income, 0) = 0 THEN 'open'
                        WHEN ABS(COALESCE(g.f26as_credit, 0.00) - COALESCE(b.system_tds, 0.00)) < 5.00 THEN 'reconciled'
                        WHEN COALESCE(g.f26as_credit, 0.00) = 0.00 THEN 'open'
                        WHEN (CASE WHEN b.system_tds >= g.f26as_credit THEN (g.f26as_credit / b.system_tds) * 100 ELSE (b.system_tds / g.f26as_credit) * 100 END) >= 85.00 THEN 'likely_match'
                        WHEN (CASE WHEN b.system_tds >= g.f26as_credit THEN (g.f26as_credit / b.system_tds) * 100 ELSE (b.system_tds / g.f26as_credit) * 100 END) >= 50.00 THEN 'suggested_match'
                        ELSE 'open'
                    END) AS recon_status,
                    COALESCE(ro.remarks, 'Automated engine validation complete.') AS remarks
                FROM cust_master c
                LEFT JOIN (
                    SELECT cust_id,
                           SUM(tds_deductible) AS system_tds,
                           SUM(gross_amt) AS gross_income 
                    FROM income_records 
                    WHERE fin_year = %s
                    GROUP BY cust_id
                ) b ON c.cust_id = b.cust_id
                LEFT JOIN (
                    SELECT LOWER(TRIM(tan_of_deductor)) AS clean_tan, 
                           SUM(tds_credited) AS f26as_credit 
                    FROM form26as_records 
                    WHERE fin_year = %s
                    GROUP BY LOWER(TRIM(tan_of_deductor))
                ) g ON (LOWER(TRIM(c.tan_number)) = g.clean_tan OR LOWER(TRIM(c.pan_number)) = g.clean_tan)
                LEFT JOIN recon_overrides ro ON ro.cust_id = c.cust_id AND ro.fin_year = %s
            ) AS main_recon
            WHERE (system_tds_amt > 0 OR f26as_credit_amt > 0 OR book_credit_amt > 0)
        """
        
        if status_filter and status_filter.lower() != 'all':
            query += " AND LOWER(recon_status) = %s"
            params.append(status_filter.lower())
        if search_q and search_q.upper() != 'ALL' and search_q != '':
            query += " AND (cust_name LIKE %s OR pan_number LIKE %s OR tan_number LIKE %s)"
            search_param = f"%{search_q}%"
            params.extend([search_param, search_param, search_param])
            
        query += " ORDER BY cust_name ASC"
        page     = max(1, int(request.args.get('page', 1)))
        per_page = max(1, min(500, int(request.args.get('per_page', 100))))
        count_q  = f"SELECT COUNT(*) AS total FROM ({query}) AS _rc"
        cursor.execute(count_q, tuple(params))
        total = cursor.fetchone()['total']
        query += f" LIMIT {per_page} OFFSET {(page-1)*per_page}"
        cursor.execute(query, tuple(params))
        rows = cursor.fetchall()
        _close(cursor, conn)
        return jsonify({'data': rows, 'total': total, 'page': page, 'per_page': per_page})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

      
# ─────────────────────────────────────────────────────────────
# LIVE PRODUCTION — TDS SUMMARY REPORT MODULE
# ─────────────────────────────────────────────────────────────  
       
@app.route('/api/tds-summary', methods=['GET'])
def get_tds_summary_report():
    raw_fy = request.args.get('fin_year', '').strip()
    cust_id = request.args.get('cust_id', '').strip()
    section_code = request.args.get('section', '').strip()
    
    fy = raw_fy.upper().replace('FY', '').replace(' ', '').strip()
    if not fy:
        return jsonify({'error': 'Missing required parameter: fin_year'}), 400
        
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        
        query = """
            SELECT 
                CONCAT(c.cust_id, '_', COALESCE(sec_agg.final_section, '—')) AS summary_id,
                c.cust_id,
                c.cust_name,
                c.pan_number,
                c.tan_number,
                COALESCE(sec_agg.final_section, '—') AS tds_section,
                COALESCE(b.total_taxable, 0.00) AS total_taxable_amt,
                COALESCE(b.total_deductible, 0.00) AS total_tds_deductible,
                COALESCE(g.total_credited, 0.00) AS total_f26as_credit,
                (COALESCE(b.total_deductible, 0.00) - COALESCE(g.total_credited, 0.00)) AS variance_amt,
                COALESCE(b.inv_count, 0) AS invoice_count
            FROM cust_master c
            
            LEFT JOIN (
                SELECT c_name, MAX(sec_code) AS final_section
                FROM (
                    SELECT cust_name AS c_name, tds_section AS sec_code FROM income_records WHERE fin_year = %s
                    UNION
                    SELECT deductor_name AS c_name, section_code AS sec_code FROM form26as_records WHERE fin_year = %s
                    UNION
                    SELECT cm.cust_name AS c_name, fr.section_code AS sec_code 
                    FROM form26as_records fr
                    JOIN file_uploads_log ul ON ul.fin_year = %s AND ul.module_name = 'form26as'
                    JOIN cust_master cm ON cm.tan_number = fr.tan_of_deductor
                    WHERE fr.fin_year = %s
                ) u_data 
                WHERE sec_code IS NOT NULL AND sec_code != '' AND sec_code != '—'
                GROUP BY c_name
            ) sec_agg ON LOWER(TRIM(c.cust_name)) = LOWER(TRIM(sec_agg.c_name))
                      
            LEFT JOIN (
                SELECT cust_id,
                       SUM(taxable_amt) AS total_taxable,
                       SUM(tds_deductible) AS total_deductible,
                       COUNT(*) AS inv_count
                FROM income_records 
                WHERE fin_year = %s
                GROUP BY cust_id
            ) b ON c.cust_id = b.cust_id
            
            LEFT JOIN (
                SELECT tan_of_deductor,
                       SUM(tds_credited) AS total_credited
                FROM form26as_records 
                WHERE fin_year = %s
                GROUP BY tan_of_deductor
            ) g ON c.tan_number = g.tan_of_deductor
            
            WHERE (COALESCE(b.total_taxable, 0) > 0 OR COALESCE(b.total_deductible, 0) > 0 OR COALESCE(g.total_credited, 0) > 0)
        """
        
        params = [fy, fy, fy, fy, fy, fy]
        
        if cust_id and cust_id.upper() != 'ALL' and cust_id != '':
            if cust_id.isdigit():
                query += " AND c.cust_id = %s"
                params.append(int(cust_id))
            else:
                query += " AND (LOWER(c.cust_name) = %s OR LOWER(c.cust_name) LIKE %s)"
                params.extend([cust_id.lower(), f"%{cust_id.lower()}%"])
                
        if section_code and section_code.upper() != 'ALL' and section_code != '':
            query += " AND sec_agg.final_section = %s"
            params.append(section_code.upper())
            
        query += " ORDER BY c.cust_name ASC"
        page     = max(1, int(request.args.get('page', 1)))
        per_page = max(1, min(500, int(request.args.get('per_page', 100))))
        count_q  = f"SELECT COUNT(*) AS total FROM ({query}) AS _tc"
        cursor.execute(count_q, tuple(params))
        total = cursor.fetchone()['total']
        query += f" LIMIT {per_page} OFFSET {(page-1)*per_page}"
        cursor.execute(query, tuple(params))
        rows = cursor.fetchall()
        _close(cursor, conn)
        return jsonify({'data': rows, 'total': total, 'page': page, 'per_page': per_page})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

        

       
# ───────────────────────────────────────────────────────────
# LIVE PRODUCTION — AUDIT TRAIL LOGISTICS CONTROLLER
# ─────────────────────────────────────────────────────────────

@app.route('/api/audit-logs', methods=['GET'])
def get_system_audit_logs():
    # 1. Capture text search keywords, dropdown items, and date streams from your front-end form
    search_q     = request.args.get('search', '').strip()
    action_val   = request.args.get('action', '').strip()
    module_val   = request.args.get('module', '').strip()
    severity_val = request.args.get('severity', '').strip()
    date_from    = request.args.get('from', '').strip()
    date_to      = request.args.get('to', '').strip()

    # 2. Pull raw fields safely — zero percent symbols means ZERO syntax crashes
    query = """
        SELECT log_id, timestamp, user_name, action_type, 
               module_name, description, severity_level, ip_address 
        FROM audit_logs 
        WHERE 1=1
    """
    params = []

    # 3. Dynamic Text Search Input
    if search_q:
        query += " AND (LOWER(description) LIKE %s OR LOWER(user_name) LIKE %s OR LOWER(module_name) LIKE %s)"
        s_p = f"%{search_q.lower()}%"
        params.extend([s_p, s_p, s_p])

    # 4. Action Type Dropdown Filter (e.g., LOGIN, UPLOAD)
    if action_val and action_val.upper() != 'ALL' and action_val != '':
        query += " AND UPPER(TRIM(action_type)) = %s"
        params.append(action_val.upper())

    # 5. Module Name Dropdown Filter (Matches your upper_underscore entries perfectly!)
    if module_val and module_val.upper() != 'ALL' and module_val != '':
        query += " AND UPPER(TRIM(module_name)) = %s"
        params.append(module_val.upper())

    # 6. Severity Level Dropdown Filter (e.g., INFO, WARNING, CRITICAL)
    if severity_val and severity_val.upper() != 'ALL' and severity_val != '':
        query += " AND UPPER(TRIM(severity_level)) = %s"
        params.append(severity_val.upper())

    # 7. Calendar Date Pickers Boundary Filters
    if date_from:
        query += " AND DATE(timestamp) >= %s"
        params.append(date_from)

    if date_to:
        query += " AND DATE(timestamp) <= %s"
        params.append(date_to)

    query += " ORDER BY timestamp DESC"
    

    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(query, tuple(params))
        rows = cursor.fetchall()
        _close(cursor, conn)
        
        # 8. HIGH UTILITY RUNTIME TRANSLATOR: Convert native timestamps into clean display strings
        for row in rows:
            if row.get('timestamp') and isinstance(row['timestamp'], datetime):
                row['timestamp'] = row['timestamp'].strftime('%Y-%m-%d %H:%M:%S')
            elif row.get('timestamp'):
                row['timestamp'] = str(row['timestamp'])
            else:
                row['timestamp'] = '—'
                
        return jsonify({'data': rows, 'total': len(rows)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500




# ──────────────────────────────────────────────────────────────
#  RECONCILIATION — MANUAL STATUS OVERRIDE
# ──────────────────────────────────────────────────────────────


@app.route('/api/recon/records/<int:recon_id>/status', methods=['PUT'])
def update_recon_status(recon_id):
    body     = request.get_json(force=True) or {}
    status   = (body.get('recon_status') or '').strip()
    remarks  = (body.get('remarks')      or '').strip()
    fin_year = (body.get('fin_year')     or '').strip()
    
    valid    = ('reconciled','likely_match','reserved','suggested_match','not_liable','open','disabled')
    if status not in valid:
        return jsonify({'error': 'Invalid status: ' + status}), 400
        
    try:
        conn = get_db()
        cur  = conn.cursor()
        
        # 1. Upsert manual status tracking into overrides schema layout
        cur.execute(
            'INSERT INTO recon_overrides (cust_id, fin_year, manual_status, remarks) '
            'VALUES (%s,%s,%s,%s) ON DUPLICATE KEY UPDATE '
            'manual_status = VALUES(manual_status), remarks = VALUES(remarks), updated_at = NOW()',
            (recon_id, fin_year, status, remarks or None)
        )
        
        # 2. Append operational event trails into audit logging matrix
        cur.execute(
            'INSERT INTO audit_logs (action_type,module_name,description,severity_level,ip_address) '
            'VALUES (%s,%s,%s,%s,%s)',
            ('UPDATE','RECONCILIATION',
             f'Status changed to {status} | cust_id={recon_id} | FY={fin_year if fin_year else "not provided"}',
             'INFO', request.remote_addr or 'unknown')
        )
        
        # Explicit commit prevents connection-close rollbacks in MySQL driver pipelines
        conn.commit()
        
        _close(cur, conn)
        return jsonify({'recon_id': recon_id, 'recon_status': status})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


    
 
# ══════════════════════════════════════════════════════════════
#  HEALTH CHECK  — visit /api/health to test DB connection
# ══════════════════════════════════════════════════════════════

@app.route('/api/health', methods=['GET'])
def health():
    try:
        conn   = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT 1')
        cursor.fetchone()
        _close(cursor, conn)
        return jsonify({'status': 'ok', 'database': 'connected'})
    except Exception as e:
        return jsonify({'status': 'error', 'database': str(e)}), 503


# ══════════════════════════════════════════════════════════════
#  ERROR HANDLERS
# ══════════════════════════════════════════════════════════════

@app.errorhandler(404)
def not_found(e):
    return jsonify({'error': 'Endpoint not found'}), 404

@app.errorhandler(405)
def method_not_allowed(e):
    return jsonify({'error': 'Method not allowed'}), 405

@app.errorhandler(500)
def internal_error(e):
    return jsonify({'error': 'Internal server error'}), 500


# ══════════════════════════════════════════════════════════════
if __name__ == '__main__':
    app.run(
        host  = os.getenv('FLASK_HOST', '127.0.0.1'),
        port  = int(os.getenv('FLASK_PORT', '5000')),
        debug = os.getenv('FLASK_DEBUG', 'true').lower() == 'true',
    )
